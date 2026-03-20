# tests/test_export_service.py

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

from app.domain.models import ControlPoint, Project
from app.services.export_service import ExportService


@pytest.fixture
def sample_project():
    return Project(
        project_id=1,
        name="Тест",
        dxf_file="test.dxf",
        description="",
        last_control="",
        points=[
            ControlPoint(
                1, "D1", "Диаметр",
                "25.0", "+0.02", "-0.01",
                0, 0, "", "25.01",
            ),
        ],
    )


def test_export_excel(sample_project):
    svc = ExportService()
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False,
    ) as f:
        path = f.name
    try:
        svc.export_excel(sample_project, path)
        assert Path(path).exists()
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "Контроль"
        assert "Тест" in [str(c.value) for c in ws[1]]
    finally:
        Path(path).unlink(missing_ok=True)


def test_export_csv(sample_project):
    svc = ExportService()
    with tempfile.NamedTemporaryFile(
        suffix=".csv",
        delete=False,
    ) as f:
        path = f.name
    try:
        svc.export_csv(sample_project, path)
        content = Path(path).read_text(encoding="utf-8")
        assert "№" in content
        assert "D1" in content
        assert "25.01" in content
    finally:
        Path(path).unlink(missing_ok=True)

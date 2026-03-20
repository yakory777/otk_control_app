"""Экспорт данных проекта в Excel, PDF, печать."""
from __future__ import annotations

import logging
from datetime import datetime

from app.domain.models import Project

logger = logging.getLogger(__name__)

_COLUMNS = [
    "№", "Название", "Тип", "Истинное",
    "Допуск +", "Допуск −",
    "Замер", "Отклонение", "Статус",
]


def _point_row(pt) -> list[str]:
    dev = pt.deviation
    dev_str = f"{dev:+.3f}" if dev is not None else "—"
    return [
        str(pt.number),
        pt.name,
        pt.kind,
        pt.true_value,
        pt.tol_plus,
        pt.tol_minus,
        pt.measured_value or "—",
        dev_str,
        pt.status,
    ]


def _build_html(
    project: Project,
    operator: str = "",
) -> str:
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    rows = ""
    for pt in project.points:
        cells = "".join(
            f"<td>{v}</td>" for v in _point_row(pt)
        )
        rows += f"<tr>{cells}</tr>\n"
    header = "".join(f"<th>{c}</th>" for c in _COLUMNS)
    op_line = (
        f" &nbsp;|&nbsp; Оператор: {operator}"
        if operator
        else ""
    )
    return f"""
    <html><head><meta charset="utf-8">
    <style>
      body {{ font-family: Arial, sans-serif;
             font-size: 10pt; }}
      h1 {{ font-size: 14pt; }}
      h2 {{ font-size: 11pt; color: #555; }}
      table {{ border-collapse: collapse;
               width: 100%; margin-top: 8px; }}
      th, td {{ border: 1px solid #999;
                padding: 4px 6px; text-align: left; }}
      th {{ background: #f0f0f0; font-size: 9pt; }}
      td {{ font-size: 9pt; }}
    </style></head><body>
    <h1>Протокол контроля</h1>
    <h2>Проект: {project.name}</h2>
    <p>Дата: {now} &nbsp;|&nbsp;
       DXF: {project.dxf_file or '—'}{op_line}</p>
    <table>
      <thead><tr>{header}</tr></thead>
      <tbody>{rows}</tbody>
    </table>
    </body></html>
    """


class ExportService:
    """Экспорт проекта в Excel / PDF / печать."""

    def __init__(self, operator: str = "") -> None:
        self._operator = operator

    def export_excel(
        self,
        project: Project,
        path: str,
    ) -> None:
        """Сохранить контрольные точки в .xlsx."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Контроль"

        ws.append(["Проект:", project.name])
        ws.append([
            "DXF:", project.dxf_file or "—",
        ])
        ws.append([
            "Дата:",
            datetime.now().strftime("%d.%m.%Y %H:%M"),
        ])
        if self._operator:
            ws.append(["Оператор:", self._operator])
        ws.append([])

        ws.append(_COLUMNS)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for pt in project.points:
            ws.append(_point_row(pt))

        ok = sum(1 for pt in project.points if pt.status == "Норма")
        warn = sum(1 for pt in project.points if pt.status == "Внимание")
        fail = sum(1 for pt in project.points if pt.status == "Брак")
        ws.append([])
        ws.append([
            "Итого:",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f"Норма: {ok}, Внимание: {warn}, Брак: {fail}",
        ])

        for col in ws.columns:
            max_len = max(
                len(str(c.value or "")) for c in col
            )
            ws.column_dimensions[
                col[0].column_letter
            ].width = min(max_len + 4, 30)

        wb.save(path)
        logger.info("Excel экспорт: %s", path)

    def export_pdf(
        self,
        project: Project,
        path: str,
    ) -> None:
        """Сохранить отчёт в PDF через QPrinter."""
        from PySide6.QtCore import QMarginsF
        from PySide6.QtGui import (
            QPageLayout,
            QPageSize,
            QTextDocument,
        )
        from PySide6.QtPrintSupport import QPrinter

        html = _build_html(project, self._operator)
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(
            QPrinter.OutputFormat.PdfFormat,
        )
        printer.setOutputFileName(path)
        printer.setPageLayout(QPageLayout(
            QPageSize(QPageSize.PageSizeId.A4),
            QPageLayout.Orientation.Portrait,
            QMarginsF(15, 15, 15, 15),
        ))
        doc = QTextDocument()
        doc.setHtml(html)
        doc.print_(printer)
        logger.info("PDF экспорт: %s", path)

    def print_report(
        self,
        project: Project,
        parent=None,
    ) -> None:
        """Открыть диалог печати."""
        from PySide6.QtGui import QTextDocument
        from PySide6.QtPrintSupport import (
            QPrintDialog,
            QPrinter,
        )

        html = _build_html(project, self._operator)
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        dlg = QPrintDialog(printer, parent)
        if dlg.exec() != QPrintDialog.DialogCode.Accepted:
            return
        doc = QTextDocument()
        doc.setHtml(html)
        doc.print_(printer)
        logger.info("Печать отчёта")

    def export_csv(
        self,
        project: Project,
        path: str,
    ) -> None:
        """Сохранить контрольные точки в CSV."""
        import csv
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(_COLUMNS)
            for pt in project.points:
                w.writerow(_point_row(pt))
        logger.info("CSV экспорт: %s", path)
